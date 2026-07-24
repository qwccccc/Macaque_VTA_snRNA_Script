from openpyxl import load_workbook


workbook_path = "/mnt/90-connectome/Personal/CQW/midbrain/upload.xlsx"
workbook = load_workbook(workbook_path, read_only=True, data_only=True)

print("sheets:", workbook.sheetnames)
for worksheet in workbook.worksheets:
    print(f"sheet: {worksheet.title}, rows={worksheet.max_row}, columns={worksheet.max_column}")
    rows = worksheet.iter_rows(values_only=True)
    print("header:", next(rows))
    for _, row in zip(range(5), rows):
        print(row)
